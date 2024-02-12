import openpyxl
from openpyxl import Workbook
import os
import shutil
import pandas as pd
from abc import ABC, abstractmethod


#Crea una clase abstracta para las reservas
class book(ABC):
    def __init__(self, fecha, factura, nombre, NIF, Base_1, Cuota_1, total, domicilio, cod_postal, pais, CL, observaciones, cuenta_contable, tipo_Sii):
        self.fecha = fecha
        self.factura = factura
        self.nombre = nombre
        self.NIF = NIF
        self.Base_1 = Base_1
        self.Cuota_1 = Cuota_1
        self.total = total
        self.domicilio = domicilio
        self.cod_postal = cod_postal
        self.pais = pais
        self.CL = CL
        self.observaciones = observaciones
        self.cuenta_contable = cuenta_contable
        self.tipo_Sii = tipo_Sii

    @abstractmethod
    def descripcion(self):
        pass
    def __str__(self):
        return f"Fecha: {self.fecha}\nFactura: {self.factura}\nNombre: {self.nombre}\nNIF: {self.NIF}\nBase: {self.Base_1}\nCuota: {self.Cuota_1}\nTotal: {self.total}\nDomicilio: {self.domicilio}\nCódigo postal: {self.cod_postal}\nPaís: {self.pais}\nCL: {self.CL}\nObservaciones: {self.observaciones}\nCuenta contable: {self.cuenta_contable}\nTipo SII: {self.tipo_Sii}\n__________________________\n"


#Crea una clase para cada una de las distintas plataformas hereadando de la clase abstracta book
class booking(book):
    def __init__(self, fecha, factura, nombre, NIF, Base_1, Cuota_1, total, domicilio, cod_postal, pais, CL, observaciones, cuenta_contable, tipo_Sii):
        super().__init__(fecha, factura, nombre, NIF, Base_1, Cuota_1, total, domicilio, cod_postal, pais, CL, observaciones, cuenta_contable, tipo_Sii)
    def descripcion(self):
        return "Factura Booking"

class airbnb(book):
    def __init__(self, fecha, factura, nombre, NIF, Base_1, Cuota_1, total, domicilio, cod_postal, pais, CL, observaciones, cuenta_contable, tipo_Sii):
        super().__init__(fecha, factura, nombre, NIF, Base_1, Cuota_1, total, domicilio, cod_postal, pais, CL, observaciones, cuenta_contable, tipo_Sii)
    def descripcion(self):
        return "Factura Airbnb"
    
class web(book):
    def __init__(self, fecha, factura, nombre, NIF, Base_1, Cuota_1, total, domicilio, cod_postal, pais, CL, observaciones, cuenta_contable, tipo_Sii):
        super().__init__(fecha, factura, nombre, NIF, Base_1, Cuota_1, total, domicilio, cod_postal, pais, CL, observaciones, cuenta_contable, tipo_Sii)
    def descripcion(self):
        return "Factura Web"
    
class error(book):
    def __init__(self, fecha, factura, nombre, NIF, Base_1, Cuota_1, total, domicilio, cod_postal, pais, CL, observaciones, cuenta_contable, tipo_Sii):
        super().__init__(fecha, factura, nombre, NIF, Base_1, Cuota_1, total, domicilio, cod_postal, pais, CL, observaciones, cuenta_contable, tipo_Sii)
    def descripcion(self):
        return "Error"

class read_excel:
    def __init__(self, file_in):
        self.file_in = file_in
        self.platforms = {
        "airbnb": airbnb,
        "booking": booking,
        "mara": web
        }

    def read_excel(self):
        wb = openpyxl.load_workbook(self.file_in)
        hoja = wb.active
        last = hoja.max_row
        
        reservas = []

        for i in range(1,last):
            fecha = str(hoja.cell(i+1,5).value)
            factura = "MBA24/" + str(hoja.cell(i+1,4).value)
            nombre = hoja.cell(i+1,14).value
            NIF = hoja.cell(i+1,15).value
            Base_1 = hoja.cell(i+1,18).value
            Cuota_1 = hoja.cell(i+1,19).value
            total = hoja.cell(i+1,20).value
            domicilio = hoja.cell(i+1,11).value
            cod_postal = hoja.cell(i+1,12).value
            pais = hoja.cell(i+1,10).value
            CL = 2
            observaciones = hoja.cell(i+1,26).value
            tipo_Sii = 1

            #Intenta crear las reservas. Si hay carácteres fantasmas continua a la siguiente iteración
            try:
                obs=observaciones.lower()
                #Comprueba en la lista de plataformas si coincide con las observaciones
                #Esto se puede optimizar??
                for i in self.platforms:
                    if i in obs:
                        #Hace "plataformas[i]" porque i es un string del diccionario e platforms[i] es el objeto
                        reservas.append(self.platforms[i](fecha, factura, nombre, NIF, Base_1, Cuota_1, total, domicilio, cod_postal, pais, CL, observaciones, 430000, tipo_Sii))
            except AttributeError:
                continue

        wb.close()
        return reservas

class write_excel:

    def __init__(self, reservas):
        self.reservas = reservas

    def empty_out(self):
        try:
            shutil.rmtree("out")
        except:
            pass
        os.mkdir("out")
    
    def conv(self, file):
        df = pd.read_excel(file, engine='openpyxl')
        file_new = file[:-5] + ".xls"
        with pd.ExcelWriter(file_new, engine='openpyxl') as writer:
            df.to_excel(writer, index=False)
        os.remove(file)
        print(f"Archivo {file_new} convertido a .xls")

    def write (self):
        self.empty_out()
        for i in self.reservas:
            clase = str(type(i).__name__)               
            file_out = "out/" + clase + ".xlsx"

            try:
                wb = openpyxl.load_workbook(file_out)
            except FileNotFoundError:
                wb = Workbook()

            hoja = wb.active
            last = hoja.max_row

            hoja.cell(last+1,1).value = i.fecha
            hoja.cell(last+1,2).value = i.factura
            hoja.cell(last+1,3).value = i.nombre
            hoja.cell(last+1,4).value = i.NIF
            hoja.cell(last+1,5).value = i.Base_1
            hoja.cell(last+1,7).value = i.Cuota_1
            hoja.cell(last+1,8).value = i.total
            hoja.cell(last+1,9).value = i.domicilio
            hoja.cell(last+1,10).value = i.cod_postal
            hoja.cell(last+1,11).value = i.pais
            hoja.cell(last+1,12).value = i.CL
            hoja.cell(last+1,14).value = i.observaciones
            hoja.cell(last+1,15).value = i.cuenta_contable
            hoja.cell(last+1,16).value = i.tipo_Sii
            last += 1
            wb.save(file_out)
            wb.close()
            print(f"Factura {i.factura} añadida a archivo {file_out}")
        
        for element in os.listdir("out"):
            self.conv("out/"+element)




excel = read_excel("in1.xlsx").read_excel()
write = write_excel(excel).write()
